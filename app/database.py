import sqlite3
import os
import csv
import re
from difflib import SequenceMatcher
from flask import g
from .config import Config
from .utils.text_utils import clean_text, norm_state, norm_org
from .utils.passwords import hash_password
from .utils.data_parse import (
    parse_meta_from_file,
    detect_status,
    detect_year,
    detect_school,
    detect_chapter,
    detect_chapter_id,
    detect_notes,
    parse_location,
)

CRM_SCHEMA_VERSION = 2

def get_connection() -> sqlite3.Connection:
    if 'db' not in g:
        g.db = sqlite3.connect(Config.DB_PATH, timeout=30)
        g.db.row_factory = sqlite3.Row
    return g.db

def close_connection(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def ensure_crm_tables(conn: sqlite3.Connection) -> None:
    current_version_row = conn.execute("PRAGMA user_version").fetchone()
    current_version = int(current_version_row[0]) if current_version_row else 0
    if current_version >= CRM_SCHEMA_VERSION:
        return

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chapter_id TEXT NOT NULL,
            org TEXT,
            chapter_name TEXT,
            school TEXT,
            city TEXT,
            state TEXT,
            status TEXT DEFAULT 'prospect',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS vendor_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT,
            chapter_id TEXT NOT NULL,
            org TEXT,
            chapter_name TEXT,
            school TEXT,
            city TEXT,
            state TEXT,
            year INTEGER,
            product TEXT,
            quantity INTEGER,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS saved_views (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            filters_json TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS chapter_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chapter_id TEXT NOT NULL,
            contact_name TEXT,
            role TEXT,
            instagram TEXT,
            email TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS lead_activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            details TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS vendor_org_licenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT NOT NULL,
            org TEXT NOT NULL,
            state TEXT,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS competitors_followed (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT NOT NULL,
            competitor_vendor TEXT NOT NULL,
            is_starred INTEGER DEFAULT 1,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(vendor, competitor_vendor)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS manufacturers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            city TEXT,
            state TEXT,
            contact_email TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS brand_owners (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact_email TEXT,
            workspace_id TEXT UNIQUE NOT NULL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS manufacturer_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            manufacturer_id INTEGER,
            manufacturer_name TEXT,
            workspace_id TEXT,
            vendor TEXT NOT NULL,
            org TEXT,
            chapter_id TEXT,
            chapter_name TEXT,
            school TEXT,
            city TEXT,
            state TEXT,
            year INTEGER,
            order_type TEXT,
            quantity INTEGER,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS manufacturer_brand_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            manufacturer_workspace_id TEXT NOT NULL,
            brand_owner_workspace_id TEXT NOT NULL,
            brand_owner_name TEXT,
            linked_by_user_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(manufacturer_workspace_id, brand_owner_workspace_id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            account_name TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            name TEXT NOT NULL,
            chapter_id TEXT,
            vendor_id INTEGER,
            connection TEXT,
            status TEXT DEFAULT 'lead',
            notes TEXT,
            last_contact_at TEXT,
            follow_up_date TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id TEXT,
            details TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crm_contact_id INTEGER,
            to_email TEXT NOT NULL,
            subject TEXT NOT NULL,
            body TEXT NOT NULL,
            status TEXT DEFAULT 'queued',
            sent_at TEXT,
            error TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crm_contact_id INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_by_user_id INTEGER,
            workspace_id TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crm_contact_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            due_date TEXT,
            status TEXT DEFAULT 'open',
            priority TEXT DEFAULT 'normal',
            created_by_user_id INTEGER,
            completed_at TEXT,
            workspace_id TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crm_contact_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            detail TEXT,
            created_by_user_id INTEGER,
            workspace_id TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            workspace_id TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(workspace_id, name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_contact_tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            crm_contact_id INTEGER NOT NULL,
            crm_tag_id INTEGER NOT NULL,
            workspace_id TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(workspace_id, crm_contact_id, crm_tag_id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_stage_titles (
            workspace_id TEXT NOT NULL,
            stage_key TEXT NOT NULL,
            title TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (workspace_id, stage_key)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            invite_code TEXT UNIQUE NOT NULL,
            workspace_id TEXT UNIQUE NOT NULL,
            owner_user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS team_invites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL,
            invite_code TEXT UNIQUE NOT NULL,
            max_uses INTEGER DEFAULT 1,
            uses INTEGER DEFAULT 0,
            expires_at TEXT,
            is_active INTEGER DEFAULT 1,
            created_by_user_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS feedback_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            workspace_id TEXT,
            page_url TEXT,
            page_title TEXT,
            message TEXT,
            image_path TEXT,
            image_name TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ops_order_brand_access (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            manufacturer_workspace_id TEXT NOT NULL,
            brand_owner_workspace_id TEXT NOT NULL,
            granted_by_user_id INTEGER,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(order_id, brand_owner_workspace_id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_research_prompts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            workspace_id TEXT NOT NULL DEFAULT '',
            category TEXT NOT NULL,
            slot_index INTEGER NOT NULL,
            label TEXT NOT NULL,
            prompt_text TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, workspace_id, category, slot_index)
        )
        """
    )

    vendor_order_columns = {row[1] for row in conn.execute("PRAGMA table_info(vendor_orders)").fetchall()}
    if "order_type" not in vendor_order_columns:
        conn.execute("ALTER TABLE vendor_orders ADD COLUMN order_type TEXT")
    if "workspace_id" not in vendor_order_columns:
        conn.execute("ALTER TABLE vendor_orders ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in vendor_order_columns:
        conn.execute("ALTER TABLE vendor_orders ADD COLUMN manufacturer_id INTEGER")
    lead_columns = {row[1] for row in conn.execute("PRAGMA table_info(leads)").fetchall()}
    if "follow_up_date" not in lead_columns:
        conn.execute("ALTER TABLE leads ADD COLUMN follow_up_date TEXT")
    if "workspace_id" not in lead_columns:
        conn.execute("ALTER TABLE leads ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in lead_columns:
        conn.execute("ALTER TABLE leads ADD COLUMN manufacturer_id INTEGER")
    users_columns = {row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
    if "account_name" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN account_name TEXT")
    if "role" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN role TEXT")
    if "manufacturer_id" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN manufacturer_id INTEGER")
    if "brand_owner_id" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN brand_owner_id INTEGER")
    if "workspace_id" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN workspace_id TEXT")
    if "account_type" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN account_type TEXT")
    if "security_question" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN security_question TEXT")
    if "security_answer_hash" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN security_answer_hash TEXT")
    if "team_id" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN team_id INTEGER")
    if "team_role" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN team_role TEXT")
    if "daily_outreach_target" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN daily_outreach_target INTEGER")
    if "google_id" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN google_id TEXT")
    if "email" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if "first_name" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN first_name TEXT")
    if "last_name" not in users_columns:
        conn.execute("ALTER TABLE users ADD COLUMN last_name TEXT")
    conn.execute("UPDATE users SET account_type='manufacturer' WHERE trim(coalesce(account_type,''))=''")
    crm_contact_columns = {row[1] for row in conn.execute("PRAGMA table_info(crm_contacts)").fetchall()}
    if "workspace_id" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN manufacturer_id INTEGER")
    if "priority" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN priority TEXT DEFAULT 'normal'")
    if "value_estimate" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN value_estimate REAL")
    if "expected_close_date" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN expected_close_date TEXT")
    if "updated_at" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN updated_at TEXT")
    if "created_by_user_id" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN created_by_user_id INTEGER")
    if "assigned_to_user_id" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN assigned_to_user_id INTEGER")
    if "contact_source" not in crm_contact_columns:
        conn.execute("ALTER TABLE crm_contacts ADD COLUMN contact_source TEXT")
    activities_columns = {row[1] for row in conn.execute("PRAGMA table_info(activities)").fetchall()}
    if "workspace_id" not in activities_columns:
        conn.execute("ALTER TABLE activities ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in activities_columns:
        conn.execute("ALTER TABLE activities ADD COLUMN manufacturer_id INTEGER")
    contacts_columns = {row[1] for row in conn.execute("PRAGMA table_info(chapter_contacts)").fetchall()}
    if "workspace_id" not in contacts_columns:
        conn.execute("ALTER TABLE chapter_contacts ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in contacts_columns:
        conn.execute("ALTER TABLE chapter_contacts ADD COLUMN manufacturer_id INTEGER")
    messages_columns = {row[1] for row in conn.execute("PRAGMA table_info(messages)").fetchall()}
    if "workspace_id" not in messages_columns:
        conn.execute("ALTER TABLE messages ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in messages_columns:
        conn.execute("ALTER TABLE messages ADD COLUMN manufacturer_id INTEGER")
    saved_views_columns = {row[1] for row in conn.execute("PRAGMA table_info(saved_views)").fetchall()}
    if "workspace_id" not in saved_views_columns:
        conn.execute("ALTER TABLE saved_views ADD COLUMN workspace_id TEXT")
    if "manufacturer_id" not in saved_views_columns:
        conn.execute("ALTER TABLE saved_views ADD COLUMN manufacturer_id INTEGER")
    lead_activities_columns = {row[1] for row in conn.execute("PRAGMA table_info(lead_activities)").fetchall()}
    if "workspace_id" not in lead_activities_columns:
        conn.execute("ALTER TABLE lead_activities ADD COLUMN workspace_id TEXT")
    feedback_columns = {row[1] for row in conn.execute("PRAGMA table_info(feedback_messages)").fetchall()}
    if "user_id" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN user_id INTEGER")
    if "workspace_id" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN workspace_id TEXT")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_brand_owners_workspace ON brand_owners(workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_users_account_type ON users(account_type)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_users_username ON users(username)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_manufacturer_brand_links_mw ON manufacturer_brand_links(manufacturer_workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_manufacturer_brand_links_bw ON manufacturer_brand_links(brand_owner_workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ops_order_brand_access_bw ON ops_order_brand_access(brand_owner_workspace_id, status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ops_order_brand_access_order ON ops_order_brand_access(order_id, status)")
    if "page_url" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN page_url TEXT")
    if "page_title" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN page_title TEXT")
    if "message" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN message TEXT")
    if "image_path" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN image_path TEXT")
    if "image_name" not in feedback_columns:
        conn.execute("ALTER TABLE feedback_messages ADD COLUMN image_name TEXT")
    if "manufacturer_id" not in lead_activities_columns:
        conn.execute("ALTER TABLE lead_activities ADD COLUMN manufacturer_id INTEGER")
    crm_notes_columns = {row[1] for row in conn.execute("PRAGMA table_info(crm_notes)").fetchall()}
    if "workspace_id" not in crm_notes_columns:
        conn.execute("ALTER TABLE crm_notes ADD COLUMN workspace_id TEXT")
    crm_tasks_columns = {row[1] for row in conn.execute("PRAGMA table_info(crm_tasks)").fetchall()}
    if "workspace_id" not in crm_tasks_columns:
        conn.execute("ALTER TABLE crm_tasks ADD COLUMN workspace_id TEXT")
    crm_activities_columns = {row[1] for row in conn.execute("PRAGMA table_info(crm_activities)").fetchall()}
    if "workspace_id" not in crm_activities_columns:
        conn.execute("ALTER TABLE crm_activities ADD COLUMN workspace_id TEXT")
    if "source" not in crm_activities_columns:
        conn.execute("ALTER TABLE crm_activities ADD COLUMN source TEXT")
    research_prompt_columns = {row[1] for row in conn.execute("PRAGMA table_info(user_research_prompts)").fetchall()}
    if "workspace_id" not in research_prompt_columns:
        conn.execute("ALTER TABLE user_research_prompts ADD COLUMN workspace_id TEXT")
    if "updated_at" not in research_prompt_columns:
        conn.execute("ALTER TABLE user_research_prompts ADD COLUMN updated_at TEXT")

    # Ensure every user has a stable workspace_id to preserve CRM ownership across sessions.
    missing_ws_users = conn.execute(
        "SELECT id, username, account_name FROM users WHERE trim(coalesce(workspace_id,''))=''"
    ).fetchall()
    for row in missing_ws_users:
        conn.execute(
            "UPDATE users SET workspace_id=? WHERE id=?",
            (
                derive_workspace_id(
                    clean_text(row["account_name"]),
                    clean_text(row["username"]),
                    int(row["id"]),
                ),
                int(row["id"]),
            ),
        )

    if current_version < 2:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS user_research_prompts_workspace_scoped (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                workspace_id TEXT NOT NULL DEFAULT '',
                category TEXT NOT NULL,
                slot_index INTEGER NOT NULL,
                label TEXT NOT NULL,
                prompt_text TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(user_id, workspace_id, category, slot_index)
            )
            """
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO user_research_prompts_workspace_scoped(
                user_id,
                workspace_id,
                category,
                slot_index,
                label,
                prompt_text,
                created_at,
                updated_at
            )
            SELECT
                urp.user_id,
                coalesce(
                    nullif(trim(urp.workspace_id), ''),
                    (
                        SELECT nullif(trim(u.workspace_id), '')
                        FROM users u
                        WHERE u.id=urp.user_id
                    ),
                    ''
                ) AS workspace_id,
                urp.category,
                urp.slot_index,
                urp.label,
                urp.prompt_text,
                coalesce(urp.created_at, CURRENT_TIMESTAMP),
                coalesce(urp.updated_at, urp.created_at, CURRENT_TIMESTAMP)
            FROM user_research_prompts urp
            ORDER BY coalesce(urp.updated_at, urp.created_at, CURRENT_TIMESTAMP) ASC, urp.id ASC
            """
        )
        conn.execute("DROP TABLE user_research_prompts")
        conn.execute("ALTER TABLE user_research_prompts_workspace_scoped RENAME TO user_research_prompts")

    # Backfill workspace on activity rows keyed by user_id where possible.
    if "workspace_id" in activities_columns:
        conn.execute(
            """
            UPDATE activities
            SET workspace_id=(
                SELECT u.workspace_id FROM users u WHERE u.id=activities.user_id
            )
            WHERE trim(coalesce(activities.workspace_id,''))=''
              AND activities.user_id IS NOT NULL
              AND EXISTS(SELECT 1 FROM users u WHERE u.id=activities.user_id)
            """
        )

    # Backfill workspace via manufacturer mapping for legacy rows.
    users_cols_after = {row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
    has_user_mid = "manufacturer_id" in users_cols_after
    manufacturer_workspace = {}
    if has_user_mid:
        mid_rows = conn.execute(
            """
            SELECT manufacturer_id, workspace_id
            FROM users
            WHERE manufacturer_id IS NOT NULL
              AND manufacturer_id > 0
              AND trim(coalesce(workspace_id,''))<>''
            ORDER BY id ASC
            """
        ).fetchall()
        for row in mid_rows:
            mid = int(row["manufacturer_id"] or 0)
            if mid <= 0 or mid in manufacturer_workspace:
                continue
            manufacturer_workspace[mid] = clean_text(row["workspace_id"])

    table_defs = [
        ("crm_contacts", "manufacturer_id"),
        ("vendor_orders", "manufacturer_id"),
        ("leads", "manufacturer_id"),
        ("messages", "manufacturer_id"),
        ("activities", "manufacturer_id"),
    ]
    for table_name, mid_col in table_defs:
        cols = {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}
        if "workspace_id" not in cols or mid_col not in cols:
            continue
        for mid, ws in manufacturer_workspace.items():
            if not ws:
                continue
            conn.execute(
                f"""
                UPDATE {table_name}
                SET workspace_id=?
                WHERE trim(coalesce(workspace_id,''))=''
                  AND {mid_col}=?
                """,
                (ws, int(mid)),
            )

    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_orders_workspace ON vendor_orders(workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_orders_workspace_vendor ON vendor_orders(workspace_id, vendor)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_orders_workspace_chapter ON vendor_orders(workspace_id, chapter_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_contacts_workspace ON crm_contacts(workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_contacts_workspace_owner ON crm_contacts(workspace_id, created_by_user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_contacts_workspace_type_status ON crm_contacts(workspace_id, type, status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_contacts_workspace_followup ON crm_contacts(workspace_id, follow_up_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_team_invites_team ON team_invites(team_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_team_invites_code ON team_invites(invite_code)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_activities_workspace ON activities(workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_activities_workspace_created ON activities(workspace_id, created_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_workspace ON leads(workspace_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_tasks_workspace_due_status ON crm_tasks(workspace_id, due_date, status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_notes_workspace_contact ON crm_notes(workspace_id, crm_contact_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_activities_workspace_contact ON crm_activities(workspace_id, crm_contact_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_activities_workspace_action ON crm_activities(workspace_id, action, created_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_crm_contact_tags_workspace_contact ON crm_contact_tags(workspace_id, crm_contact_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_research_prompts_user_workspace_category ON user_research_prompts(user_id, workspace_id, category)")
    conn.execute(f"PRAGMA user_version = {CRM_SCHEMA_VERSION}")
    conn.commit()

def derive_workspace_id(account_name: str = "", username: str = "", user_id: int = 0) -> str:
    base = clean_text(account_name) or clean_text(username)
    slug = re.sub(r"[^a-z0-9]+", "-", base.lower()).strip("-")
    if slug:
        return f"ws-{slug}"
    if int(user_id or 0) > 0:
        return f"user-{int(user_id)}"
    return "ws-default"

def log_lead_activity(
    conn: sqlite3.Connection,
    lead_id: int,
    action: str,
    details: str = "",
    workspace_id: str = "",
) -> None:
    conn.execute(
        "INSERT INTO lead_activities (lead_id, action, details, workspace_id) VALUES (?, ?, ?, ?)",
        (int(lead_id), clean_text(action), clean_text(details), clean_text(workspace_id)),
    )

def ensure_default_users(conn: sqlite3.Connection) -> None:
    defaults = [
        ("demo", "demo123", "Demo Account"),
    ]
    users_columns = {row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
    for username, password, account_name in defaults:
        existing = conn.execute("SELECT id FROM users WHERE lower(username)=lower(?)", (username,)).fetchone()
        if existing:
            if "workspace_id" in users_columns:
                user_row = conn.execute(
                    "SELECT id, username, account_name, workspace_id FROM users WHERE id=?",
                    (int(existing["id"]),),
                ).fetchone()
                if user_row and not clean_text(user_row["workspace_id"]):
                    conn.execute(
                        "UPDATE users SET workspace_id=? WHERE id=?",
                        (
                            derive_workspace_id(
                                clean_text(user_row["account_name"]),
                                clean_text(user_row["username"]),
                                int(user_row["id"]),
                            ),
                            int(user_row["id"]),
                        ),
                    )
            continue

        workspace_id = derive_workspace_id(account_name, username)
        if "role" in users_columns and "workspace_id" in users_columns:
            conn.execute(
                "INSERT INTO users(username, password_hash, account_name, role, workspace_id) VALUES(?, ?, ?, 'admin', ?)",
                (username, hash_password(password), account_name, workspace_id),
            )
        else:
            conn.execute(
                "INSERT INTO users(username, password_hash, account_name) VALUES(?, ?, ?)",
                (username, hash_password(password), account_name),
            )
    conn.commit()

def log_activity(
    conn: sqlite3.Connection,
    user_id: int,
    action: str,
    entity_type: str = "",
    entity_id: str = "",
    details: str = "",
    workspace_id: str = "",
    manufacturer_id: int = 0,
) -> None:
    uid = int(user_id or 0) or None
    mid = int(manufacturer_id or 0)
    ws = clean_text(workspace_id)

    users_cols = {row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
    if uid:
        user_row = conn.execute(
            "SELECT manufacturer_id, workspace_id, account_name, username FROM users WHERE id=?",
            (uid,),
        ).fetchone()
        if user_row is not None:
            if "manufacturer_id" in users_cols and not mid:
                mid = int(user_row["manufacturer_id"] or 0)
            if not ws:
                ws = clean_text(user_row["workspace_id"]) or derive_workspace_id(
                    clean_text(user_row["account_name"]),
                    clean_text(user_row["username"]),
                    uid,
                )
    if not ws:
        ws = derive_workspace_id(user_id=uid or 0)
    if mid < 0:
        mid = 0

    activity_cols = {row[1] for row in conn.execute("PRAGMA table_info(activities)").fetchall()}
    cols = []
    vals = []
    if "manufacturer_id" in activity_cols:
        cols.append("manufacturer_id")
        vals.append(mid)
    if "workspace_id" in activity_cols:
        cols.append("workspace_id")
        vals.append(ws)
    if "user_id" in activity_cols:
        cols.append("user_id")
        vals.append(uid)
    cols.extend(["action", "entity_type", "entity_id", "details"])
    vals.extend(
        [
            clean_text(action),
            clean_text(entity_type),
            clean_text(entity_id),
            clean_text(details),
        ]
    )
    placeholders = ",".join("?" for _ in cols)
    conn.execute(f"INSERT INTO activities({','.join(cols)}) VALUES({placeholders})", tuple(vals))

def ensure_vendor_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS vendors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT,
            organization TEXT,
            registered_org_count INTEGER,
            website TEXT,
            city TEXT,
            state TEXT,
            category TEXT,
            email TEXT,
            phone TEXT,
            instagram_url TEXT,
            organization_norm TEXT,
            state_norm TEXT,
            license_label TEXT,
            is_greek_licensed INTEGER DEFAULT 0,
            is_collegiate INTEGER DEFAULT 0
        )
        """
    )

    vendor_columns = {row[1] for row in conn.execute("PRAGMA table_info(vendors)").fetchall()}
    if "phone" not in vendor_columns:
        conn.execute("ALTER TABLE vendors ADD COLUMN phone TEXT")
    if "license_label" not in vendor_columns:
        conn.execute("ALTER TABLE vendors ADD COLUMN license_label TEXT")
    if "is_greek_licensed" not in vendor_columns:
        conn.execute("ALTER TABLE vendors ADD COLUMN is_greek_licensed INTEGER DEFAULT 0")
    if "is_collegiate" not in vendor_columns:
        conn.execute("ALTER TABLE vendors ADD COLUMN is_collegiate INTEGER DEFAULT 0")
    if "instagram_url" not in vendor_columns:
        conn.execute("ALTER TABLE vendors ADD COLUMN instagram_url TEXT")

    greek_exists = os.path.exists(Config.VENDOR_CSV_PATH)
    collegiate_exists = os.path.exists(Config.COLLEGIATE_VENDOR_CSV_PATH)
    if not greek_exists and not collegiate_exists:
        return

    greek_mtime = str(os.path.getmtime(Config.VENDOR_CSV_PATH)) if greek_exists else ""
    collegiate_mtime = str(os.path.getmtime(Config.COLLEGIATE_VENDOR_CSV_PATH)) if collegiate_exists else ""
    prev_greek = conn.execute("SELECT value FROM app_meta WHERE key='vendors_csv_mtime'").fetchone()
    prev_collegiate = conn.execute("SELECT value FROM app_meta WHERE key='collegiate_vendors_csv_mtime'").fetchone()
    if prev_greek and prev_collegiate:
        if (prev_greek[0] or "") == greek_mtime and (prev_collegiate[0] or "") == collegiate_mtime:
            return

    def split_city_state(value: str) -> tuple[str, str]:
        raw = clean_text(value)
        if not raw:
            return "", ""
        if "," in raw:
            city, state = raw.rsplit(",", 1)
            return city.strip(), state.strip()
        return raw, ""

    def license_label(greek_flag: bool, collegiate_flag: bool) -> str:
        if greek_flag and collegiate_flag:
            return "Greek Licensed Holding + Collegiate Vendor"
        if greek_flag:
            return "Greek Licensed Holding"
        if collegiate_flag:
            return "Collegiate Vendor"
        return ""

    conn.execute("DELETE FROM vendors")

    vendor_flags = {}
    records = []

    if greek_exists:
        with open(Config.VENDOR_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                vendor = clean_text(row.get("Vendor"))
                organization = clean_text(row.get("Organization"))
                reg_count_raw = clean_text(row.get("RegisteredOrgCount"))
                reg_count = int(reg_count_raw) if reg_count_raw.isdigit() else None
                website = clean_text(row.get("Website"))
                city = clean_text(row.get("City"))
                state = norm_state(row.get("State")) or clean_text(row.get("State"))
                category = clean_text(row.get("Category"))
                email = clean_text(row.get("Email"))
                vendor_norm = clean_text(vendor).lower()
                if vendor_norm:
                    vendor_flags.setdefault(vendor_norm, {"greek": False, "collegiate": False})
                    vendor_flags[vendor_norm]["greek"] = True
                records.append(
                    {
                        "vendor": vendor,
                        "organization": organization,
                        "registered_org_count": reg_count,
                        "website": website,
                        "city": city,
                        "state": state,
                        "category": category,
                        "email": email,
                        "phone": "",
                        "instagram_url": "",
                        "organization_norm": norm_org(organization),
                        "state_norm": norm_state(state),
                        "vendor_norm": vendor_norm,
                    }
                )

    if collegiate_exists:
        with open(Config.COLLEGIATE_VENDOR_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                vendor = clean_text(row.get("nobull"))
                city_state_raw = clean_text(row.get("nobull 2"))
                phone = clean_text(row.get("nobull 3"))
                website_text = clean_text(row.get("nobull 4"))
                website_url = clean_text(row.get("nobull href"))
                email = clean_text(row.get("nobull 5"))
                email_url = clean_text(row.get("nobull href 2"))
                notes = clean_text(row.get("nobull 6"))
                if not any([vendor, city_state_raw, phone, website_text, website_url, email, email_url, notes]):
                    continue
                city, state_raw = split_city_state(city_state_raw)
                state = norm_state(state_raw) or clean_text(state_raw)
                website = website_url or website_text
                vendor_norm = clean_text(vendor).lower()
                if vendor_norm:
                    vendor_flags.setdefault(vendor_norm, {"greek": False, "collegiate": False})
                    vendor_flags[vendor_norm]["collegiate"] = True
                records.append(
                    {
                        "vendor": vendor,
                        "organization": "",
                        "registered_org_count": None,
                        "website": website,
                        "city": city,
                        "state": state,
                        "category": "",
                        "email": email,
                        "phone": phone,
                        "instagram_url": "",
                        "organization_norm": "",
                        "state_norm": norm_state(state),
                        "vendor_norm": vendor_norm,
                    }
                )

    merged = {}
    for rec in records:
        vendor_norm = rec.get("vendor_norm") or ""
        state_norm = clean_text(rec.get("state_norm"))
        city_norm = clean_text(rec.get("city")).lower()
        key = (vendor_norm, state_norm, city_norm)
        if key not in merged:
            merged[key] = dict(rec)
            continue
        current = merged[key]
        for field in ["organization", "website", "city", "state", "category", "email"]:
            if not clean_text(current.get(field)) and clean_text(rec.get(field)):
                current[field] = rec.get(field)
        if not clean_text(current.get("phone")) and clean_text(rec.get("phone")):
            current["phone"] = rec.get("phone")
        if current.get("registered_org_count") is None and rec.get("registered_org_count") is not None:
            current["registered_org_count"] = rec.get("registered_org_count")
        if not clean_text(current.get("organization_norm")) and clean_text(rec.get("organization_norm")):
            current["organization_norm"] = rec.get("organization_norm")
        if not clean_text(current.get("state_norm")) and clean_text(rec.get("state_norm")):
            current["state_norm"] = rec.get("state_norm")

    batch = []
    for rec in merged.values():
        vendor_norm = rec.get("vendor_norm") or ""
        flags = vendor_flags.get(vendor_norm, {"greek": False, "collegiate": False})
        greek_flag = bool(flags.get("greek"))
        collegiate_flag = bool(flags.get("collegiate"))
        batch.append(
            (
                rec.get("vendor"),
                rec.get("organization"),
                rec.get("registered_org_count"),
                rec.get("website"),
                rec.get("city"),
                rec.get("state"),
                rec.get("category"),
                rec.get("email"),
                rec.get("phone"),
                rec.get("instagram_url"),
                rec.get("organization_norm"),
                rec.get("state_norm"),
                license_label(greek_flag, collegiate_flag),
                1 if greek_flag else 0,
                1 if collegiate_flag else 0,
            )
        )

    if batch:
        conn.executemany(
            """
            INSERT INTO vendors
            (vendor, organization, registered_org_count, website, city, state, category, email, phone, instagram_url, organization_norm, state_norm,
             license_label, is_greek_licensed, is_collegiate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )

    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('vendors_csv_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (greek_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('collegiate_vendors_csv_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (collegiate_mtime,),
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_name ON vendors(vendor)")
    conn.commit()


INSTITUTION_MATCH_JOINERS = {"the", "of", "at", "and", "for"}


def _institution_match_tokens(value: object) -> list[str]:
    text = clean_text(value).lower()
    if not text:
        return []
    text = text.replace("&", " and ").replace("-", " ")
    text = re.sub(r"\buniv[.]?\b", "university", text)
    text = re.sub(r"\bst[.]?(?=\s+[a-z])", "saint", text)
    return [
        token
        for token in re.findall(r"[a-z0-9]+", text)
        if token and token not in INSTITUTION_MATCH_JOINERS
    ]


def _institution_match_key(value: object) -> str:
    return " ".join(_institution_match_tokens(value))


def _institution_name_variants(row: sqlite3.Row | dict) -> list[dict]:
    seen = set()
    variants = []
    for field in ("location_name", "alias", "parent_name"):
        raw = clean_text(row[field]) if isinstance(row, sqlite3.Row) else clean_text(row.get(field))
        if not raw:
            continue
        pieces = [raw]
        if field == "alias":
            pieces.extend([clean_text(part) for part in re.split(r"\|+|;+|\n+", raw) if clean_text(part)])
        for piece in pieces:
            key = _institution_match_key(piece)
            compact = norm_org(piece)
            if not key and not compact:
                continue
            marker = (key, compact)
            if marker in seen:
                continue
            seen.add(marker)
            variants.append(
                {
                    "raw": piece,
                    "key": key,
                    "compact": compact,
                    "tokens": set(key.split()) if key else set(),
                }
            )
    return variants


def _build_institution_match_index(inst_rows: list[sqlite3.Row]) -> dict:
    all_rows = []
    by_state = {}
    exact_lookup = {}
    key_lookup = {}

    for row in inst_rows:
        state_name = norm_state(row["state"])
        state_code = ""
        if state_name:
            for abbr, full_name in Config.STATE_ABBR.items():
                if full_name == state_name:
                    state_code = abbr
                    break
        item = {
            "id": int(row["id"]),
            "location_name": clean_text(row["location_name"]),
            "alias": clean_text(row["alias"]) if "alias" in row.keys() else "",
            "parent_name": clean_text(row["parent_name"]) if "parent_name" in row.keys() else "",
            "city": clean_text(row["city"]) if "city" in row.keys() else "",
            "state": state_name,
            "state_code": state_code,
            "variants": _institution_name_variants(row),
        }
        if not item["variants"]:
            continue
        all_rows.append(item)
        if state_code:
            by_state.setdefault(state_code, []).append(item)
        for variant in item["variants"]:
            if variant["compact"]:
                exact_lookup.setdefault((state_code, variant["compact"]), set()).add(item["id"])
                exact_lookup.setdefault(("", variant["compact"]), set()).add(item["id"])
            if variant["key"]:
                key_lookup.setdefault((state_code, variant["key"]), set()).add(item["id"])
                key_lookup.setdefault(("", variant["key"]), set()).add(item["id"])

    return {
        "all_rows": all_rows,
        "by_state": by_state,
        "exact_lookup": exact_lookup,
        "key_lookup": key_lookup,
    }


def _match_chapter_to_institution(
    school: object,
    state: object,
    city: object,
    match_index: dict,
) -> int | None:
    school_text = clean_text(school)
    school_compact = norm_org(school_text)
    school_key = _institution_match_key(school_text)
    school_tokens = set(school_key.split()) if school_key else set()
    if not school_compact and not school_key:
        return None

    state_name = norm_state(state)
    state_code = ""
    if state_name:
        for abbr, full_name in Config.STATE_ABBR.items():
            if full_name == state_name:
                state_code = abbr
                break
    city_text = clean_text(city).lower()

    for lookup, probe in (
        (match_index["exact_lookup"], school_compact),
        (match_index["key_lookup"], school_key),
    ):
        if not probe:
            continue
        state_matches = lookup.get((state_code, probe), set()) if state_code else set()
        if len(state_matches) == 1:
            return next(iter(state_matches))
        global_matches = lookup.get(("", probe), set())
        if len(global_matches) == 1:
            return next(iter(global_matches))

    candidates = match_index["by_state"].get(state_code) if state_code else None
    if not candidates:
        candidates = match_index["all_rows"]

    scored = []
    for candidate in candidates:
        best_variant_score = 0.0
        for variant in candidate["variants"]:
            if school_key and variant["key"] and school_key == variant["key"]:
                best_variant_score = max(best_variant_score, 0.985)
            if school_tokens and variant["tokens"]:
                overlap = school_tokens & variant["tokens"]
                if len(overlap) >= 2 and (
                    school_tokens.issubset(variant["tokens"]) or variant["tokens"].issubset(school_tokens)
                ):
                    best_variant_score = max(best_variant_score, 0.915)
            compact_ratio = (
                SequenceMatcher(None, school_compact, variant["compact"]).ratio()
                if school_compact and variant["compact"]
                else 0.0
            )
            key_ratio = (
                SequenceMatcher(None, school_key, variant["key"]).ratio()
                if school_key and variant["key"]
                else 0.0
            )
            best_variant_score = max(best_variant_score, compact_ratio, key_ratio)

        if best_variant_score <= 0:
            continue
        if city_text and clean_text(candidate["city"]).lower() == city_text:
            best_variant_score += 0.035
        scored.append((best_variant_score, candidate["id"]))

    if not scored:
        return None

    scored.sort(reverse=True)
    best_score, best_id = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    score_gap = best_score - second_score

    if best_score >= 0.985:
        return best_id
    if city_text and best_score >= 0.86 and score_gap >= 0.03:
        return best_id
    if best_score >= 0.92 and score_gap >= 0.025:
        return best_id
    if len(scored) == 1 and best_score >= 0.86:
        return best_id
    return None


def _refresh_chapter_institution_links(conn: sqlite3.Connection) -> int:
    chapters_info = conn.execute("PRAGMA table_info(chapters)").fetchall()
    if not chapters_info:
        return 0

    inst_rows = conn.execute(
        "SELECT id, location_name, alias, parent_name, city, state FROM institutions"
    ).fetchall()
    if not inst_rows:
        return 0
    match_index = _build_institution_match_index(inst_rows)
    chapter_rows = conn.execute("SELECT id, school, city, state, institution_id FROM chapters").fetchall()
    updates = []
    for row in chapter_rows:
        inst_id = _match_chapter_to_institution(row["school"], row["state"], row["city"], match_index)
        current_inst = int(row["institution_id"]) if row["institution_id"] is not None else None
        if inst_id and current_inst != inst_id:
            updates.append((inst_id, int(row["id"])))

    if updates:
        conn.executemany("UPDATE chapters SET institution_id=? WHERE id=?", updates)
    return len(updates)

def ensure_institutions_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS institutions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dapip_id TEXT,
            ope_id TEXT,
            ipeds_unit_ids TEXT,
            location_name TEXT,
            parent_name TEXT,
            parent_dapip_id TEXT,
            location_type TEXT,
            address TEXT,
            street TEXT,
            city TEXT,
            state TEXT,
            zip TEXT,
            general_phone TEXT,
            admin_name TEXT,
            admin_phone TEXT,
            admin_email TEXT,
            fax TEXT,
            update_date TEXT,
            state_norm TEXT,
            location_name_norm TEXT,
            parent_name_norm TEXT
        )
        """
    )

    columns = {row[1] for row in conn.execute("PRAGMA table_info(institutions)").fetchall()}
    add_cols = [
        ("street", "TEXT"),
        ("city", "TEXT"),
        ("state", "TEXT"),
        ("zip", "TEXT"),
        ("state_norm", "TEXT"),
        ("location_name_norm", "TEXT"),
        ("parent_name_norm", "TEXT"),
        ("institution_id", "TEXT"),
        ("alias", "TEXT"),
        ("zip_five_digit", "TEXT"),
        ("fips_state_code", "TEXT"),
        ("telephone", "TEXT"),
        ("ein", "TEXT"),
        ("website", "TEXT"),
        ("unitid", "TEXT"),
        ("students_total", "INTEGER"),
        ("dorm_capacity", "INTEGER"),
        ("acceptance_rate", "REAL"),
        ("institution_level", "TEXT"),
        ("control", "TEXT"),
        ("highest_offering", "TEXT"),
        ("ug_offering", "TEXT"),
        ("grad_offering", "TEXT"),
        ("degree_granting_status", "TEXT"),
        ("locale", "TEXT"),
        ("public_status", "TEXT"),
        ("post_secondary_status", "TEXT"),
        ("fips_county_code", "TEXT"),
        ("county", "TEXT"),
        ("congressional_district", "TEXT"),
        ("longitude", "REAL"),
        ("latitude", "REAL"),
    ]
    for col, ctype in add_cols:
        if col not in columns:
            conn.execute(f"ALTER TABLE institutions ADD COLUMN {col} {ctype}")

    acc_csv_exists = os.path.exists(Config.ACCREDITED_INSTITUTIONS_CSV_PATH)
    acc_xlsx_exists = os.path.exists(Config.ACCREDITED_INSTITUTIONS_XLSX_PATH)
    hd_exists = os.path.exists(Config.IPEDS_HD2024_PATH)
    ef_exists = os.path.exists(Config.IPEDS_EF2024A_PATH)
    ic_exists = os.path.exists(Config.IPEDS_IC2024_PATH)
    drv_exists = os.path.exists(Config.IPEDS_DRVADM2024_PATH)
    if not acc_csv_exists and not acc_xlsx_exists and not hd_exists and not ef_exists and not ic_exists and not drv_exists:
        return

    acc_mtime = ""
    if acc_csv_exists:
        acc_mtime = str(os.path.getmtime(Config.ACCREDITED_INSTITUTIONS_CSV_PATH))
    elif acc_xlsx_exists:
        acc_mtime = str(os.path.getmtime(Config.ACCREDITED_INSTITUTIONS_XLSX_PATH))
    hd_mtime = str(os.path.getmtime(Config.IPEDS_HD2024_PATH)) if hd_exists else ""
    ef_mtime = str(os.path.getmtime(Config.IPEDS_EF2024A_PATH)) if ef_exists else ""
    ic_mtime = str(os.path.getmtime(Config.IPEDS_IC2024_PATH)) if ic_exists else ""
    drv_mtime = str(os.path.getmtime(Config.IPEDS_DRVADM2024_PATH)) if drv_exists else ""

    prev_acc = conn.execute("SELECT value FROM app_meta WHERE key='institutions_accredited_mtime'").fetchone()
    prev_hd = conn.execute("SELECT value FROM app_meta WHERE key='institutions_ipeds_hd2024_mtime'").fetchone()
    prev_ef = conn.execute("SELECT value FROM app_meta WHERE key='institutions_ipeds_ef2024a_mtime'").fetchone()
    prev_ic = conn.execute("SELECT value FROM app_meta WHERE key='institutions_ipeds_ic2024_mtime'").fetchone()
    prev_drv = conn.execute("SELECT value FROM app_meta WHERE key='institutions_ipeds_drvadm2024_mtime'").fetchone()
    prev_link_max = conn.execute("SELECT value FROM app_meta WHERE key='chapters_link_max_id'").fetchone()
    source_tag = "ipeds_2024_combined_v4"
    prev_source = conn.execute("SELECT value FROM app_meta WHERE key='institutions_source'").fetchone()
    if prev_source and prev_acc and prev_hd and prev_ef and prev_ic and prev_drv:
        if (
            (prev_source[0] or "") == source_tag
            and (prev_acc[0] or "") == acc_mtime
            and (prev_hd[0] or "") == hd_mtime
            and (prev_ef[0] or "") == ef_mtime
            and (prev_ic[0] or "") == ic_mtime
            and (prev_drv[0] or "") == drv_mtime
        ):
            chapters_info = conn.execute("PRAGMA table_info(chapters)").fetchall()
            if chapters_info:
                current_link_max = conn.execute("SELECT COALESCE(MAX(id), 0) FROM chapters").fetchone()[0]
                current_link_max_text = str(int(current_link_max or 0))
                if prev_link_max and clean_text(prev_link_max[0]) == current_link_max_text:
                    return
                if not prev_link_max:
                    if int(current_link_max or 0) <= 100:
                        _refresh_chapter_institution_links(conn)
                    conn.execute(
                        "INSERT INTO app_meta(key, value) VALUES('chapters_link_max_id', ?) "
                        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                        (current_link_max_text,),
                    )
                    conn.commit()
                    return
                _refresh_chapter_institution_links(conn)
                conn.execute(
                    "INSERT INTO app_meta(key, value) VALUES('chapters_link_max_id', ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (current_link_max_text,),
                )
                conn.commit()
                return
            return

    def parse_address(raw: str) -> tuple[str, str, str, str]:
        raw = clean_text(raw)
        if not raw:
            return "", "", "", ""
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        street = ""
        city = ""
        state = ""
        zip_code = ""
        state_zip = ""
        if len(parts) >= 3:
            street = ", ".join(parts[:-2]).strip()
            city = parts[-2].strip()
            state_zip = parts[-1].strip()
        elif len(parts) == 2:
            street = parts[0].strip()
            state_zip = parts[1].strip()
        else:
            street = raw
        if state_zip:
            m = re.search(r"([A-Za-z]{2})\\s*(\\d{5}(?:-\\d{4})?)?", state_zip)
            if m:
                state_abbr = m.group(1).upper()
                zip_code = m.group(2) or ""
                state = norm_state(state_abbr) or state_abbr
            else:
                state = norm_state(state_zip) or state_zip
        return street, city, state, zip_code

    def name_key(name: str, state: str, city: str) -> tuple[str, str, str]:
        return (norm_org(name), norm_state(state) or clean_text(state), clean_text(city).lower())

    def split_ipeds(raw: str) -> list[str]:
        if not raw:
            return []
        parts = re.split(r"[;,\\s]+", raw)
        return [p for p in parts if p and p.isdigit()]

    def parse_int(raw: object):
        val = clean_text(raw)
        if not val:
            return None
        try:
            return int(float(val))
        except Exception:
            return None

    def parse_float(raw: object):
        val = clean_text(raw)
        if not val:
            return None
        try:
            return float(val)
        except Exception:
            return None

    def map_control(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "Public Institution",
                "2": "Private, Not-for-profit Institution",
                "3": "Private, For-profit Institution",
            }.get(val, val)
        return val

    def map_level(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "4 year",
                "2": "2 year",
                "3": "Less than 2 year",
            }.get(val, val)
        return val

    def map_degree_grant(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "Degree-granting",
                "2": "Nondegree-granting",
            }.get(val, "")
        return val

    def map_highest_offering(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "Less than 1 year",
                "2": "At least 1 year but less than 2 years",
                "3": "Associate's degree",
                "4": "At least 2 but less than 4 years",
                "5": "Bachelor's degree",
                "6": "Postbaccalaureate certificate",
                "7": "Master's degree",
                "8": "Post-master's certificate",
                "9": "Doctor's degree",
            }.get(val, "")
        return val

    def map_offer(raw: str, label: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return f"Offers {label}" if val == "1" else f"No {label}" if val == "2" else ""
        return val

    def map_locale(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "11": "City: Large (250K+)",
                "12": "City: Midsize (100-250K)",
                "13": "City: Small (<100K)",
                "21": "Suburb: Large (250K+)",
                "22": "Suburb: Midsize (100-250K)",
                "23": "Suburb: Small (<100K)",
                "31": "Town: Fringe (<10 mi)",
                "32": "Town: Distant (10-35 mi)",
                "33": "Town: Remote (35+ mi)",
                "41": "Rural: Fringe (<5 mi)",
                "42": "Rural: Distant (5-25 mi)",
                "43": "Rural: Remote (25+ mi)",
            }.get(val, "")
        return val

    def map_public_status(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "Open to public",
                "2": "Not open to public",
            }.get(val, "")
        return val

    def map_post_secondary(raw: str) -> str:
        val = clean_text(raw)
        if val.isdigit():
            return {
                "1": "Post-secondary",
                "2": "Not post-secondary",
            }.get(val, "")
        return val

    def merge_record(target: dict, updates: dict, overwrite: bool = True) -> None:
        for key, value in updates.items():
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            if overwrite or not clean_text(target.get(key)):
                target[key] = value

    def ensure_norm_fields(target: dict) -> None:
        target["state_norm"] = norm_state(target.get("state")) or clean_text(target.get("state"))
        target["location_name_norm"] = norm_org(target.get("location_name"))
        target["parent_name_norm"] = norm_org(target.get("parent_name"))

    if not prev_source or (prev_source[0] or "") != source_tag:
        conn.execute("DELETE FROM institutions")
        existing_rows = []
    else:
        existing_rows = conn.execute("SELECT * FROM institutions").fetchall()
    records = {}
    by_ope = {}
    by_ipeds = {}
    by_name = {}
    new_by_ope = {}
    new_by_ipeds = {}
    new_by_name = {}

    for row in existing_rows:
        rec = {k: row[k] for k in row.keys()}
        rec_id = int(rec["id"])
        records[rec_id] = rec
        ope_id = clean_text(rec.get("ope_id"))
        if ope_id:
            by_ope.setdefault(ope_id, rec_id)
        for ipeds in split_ipeds(clean_text(rec.get("ipeds_unit_ids"))):
            by_ipeds.setdefault(ipeds, rec_id)
        key = name_key(clean_text(rec.get("location_name")), clean_text(rec.get("state")), clean_text(rec.get("city")))
        if key[0]:
            by_name.setdefault(key, rec_id)

    def match_existing(updates: dict):
        ope_id = clean_text(updates.get("ope_id"))
        if ope_id and ope_id in by_ope:
            return ("existing", by_ope[ope_id])
        ipeds = clean_text(updates.get("institution_id"))
        if not ipeds:
            ipeds_list = split_ipeds(clean_text(updates.get("ipeds_unit_ids")))
            ipeds = ipeds_list[0] if ipeds_list else ""
        if ipeds and ipeds in by_ipeds:
            return ("existing", by_ipeds[ipeds])
        key = name_key(clean_text(updates.get("location_name")), clean_text(updates.get("state")), clean_text(updates.get("city")))
        if key[0] and key in by_name:
            return ("existing", by_name[key])
        if ope_id and ope_id in new_by_ope:
            return ("new", new_by_ope[ope_id])
        if ipeds and ipeds in new_by_ipeds:
            return ("new", new_by_ipeds[ipeds])
        if key[0] and key in new_by_name:
            return ("new", new_by_name[key])
        return (None, None)

    new_records: list[dict] = []

    def load_accredited_rows() -> list[dict]:
        if acc_csv_exists:
            with open(Config.ACCREDITED_INSTITUTIONS_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                return [dict(row) for row in reader]
        if acc_xlsx_exists:
            try:
                from openpyxl import load_workbook
            except Exception:
                return []
            wb = load_workbook(Config.ACCREDITED_INSTITUTIONS_XLSX_PATH, read_only=True, data_only=True)
            ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
            header = None
            header_row = None
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row and any(isinstance(v, str) and "INSTITUTION NAME" in v.upper() for v in row if v):
                    header = [clean_text(v) for v in row]
                    header_row = idx
                    break
            if not header or header_row is None:
                return []
            out = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not any(v is not None and str(v).strip() != "" for v in row):
                    continue
                data = {}
                for idx, col in enumerate(header):
                    if not col:
                        continue
                    data[col] = row[idx] if idx < len(row) else None
                out.append(data)
            return out
        return []

    accredited_rows = load_accredited_rows()
    for row in accredited_rows:
        updates = {
            "institution_id": clean_text(row.get("INSTITUTION ID")),
            "unitid": clean_text(row.get("INSTITUTION ID")),
            "location_name": clean_text(row.get("INSTITUTION NAME")),
            "alias": clean_text(row.get("ALIAS")),
            "address": clean_text(row.get("ADDRESS")),
            "city": clean_text(row.get("CITY")),
            "state": clean_text(row.get("STATE (ABRV)")),
            "zip": clean_text(row.get("ZIP")),
            "zip_five_digit": clean_text(row.get("ZIP_fiveDigit")),
            "fips_state_code": clean_text(row.get("FIPS STATE COD")),
            "telephone": clean_text(row.get("TELEPHONE")),
            "ein": clean_text(row.get("EIN #")),
            "ope_id": clean_text(row.get("OPE ID #")),
            "website": clean_text(row.get("WEBSITE")),
            "institution_level": clean_text(row.get("INSTITUTION LEVEL")),
            "control": clean_text(row.get("CONTROL")),
            "highest_offering": clean_text(row.get("HIGHEST OFFERING")),
            "ug_offering": clean_text(row.get("UG OFFERING")),
            "grad_offering": clean_text(row.get("GRAD OFFERING")),
            "degree_granting_status": clean_text(row.get("DEGREE-GRANTING STATUS")),
            "locale": clean_text(row.get("LOCALE")),
            "public_status": clean_text(row.get("PUBLIC STATUS")),
            "post_secondary_status": clean_text(row.get("POST SECONDARY STATUS")),
            "fips_county_code": clean_text(row.get("FIPS COUNTY CODE")),
            "county": clean_text(row.get("COUNTY")),
            "congressional_district": clean_text(row.get("CONGRESSIONAL DISTRICT")),
            "longitude": row.get("LONGITUDE"),
            "latitude": row.get("LATITUDE"),
        }
        ensure_norm_fields(updates)
        match_kind, match_id = match_existing(updates)
        if match_kind == "existing":
            merge_record(records[match_id], updates, overwrite=False)
        elif match_kind == "new":
            merge_record(new_records[match_id], updates, overwrite=False)
        else:
            new_records.append(dict(updates))
            new_index = len(new_records) - 1
            if updates.get("ope_id"):
                new_by_ope.setdefault(clean_text(updates.get("ope_id")), new_index)
            if updates.get("institution_id"):
                new_by_ipeds.setdefault(clean_text(updates.get("institution_id")), new_index)
            key = name_key(clean_text(updates.get("location_name")), clean_text(updates.get("state")), clean_text(updates.get("city")))
            if key[0]:
                new_by_name.setdefault(key, new_index)

    def load_csv_column(path: str, value_col: str, parser, row_filter=None):
        if not os.path.exists(path):
            return {}
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or "UNITID" not in reader.fieldnames or value_col not in reader.fieldnames:
                return {}
            out = {}
            for row in reader:
                if row_filter is not None and not row_filter(row):
                    continue
                unitid = clean_text(row.get("UNITID"))
                if not unitid:
                    continue
                if unitid in out:
                    continue
                val = parser(row.get(value_col))
                if val is None:
                    continue
                out[unitid] = val
            return out

    def ef_total_filter(row: dict) -> bool:
        return clean_text(row.get("EFALEVEL")) == "1" and clean_text(row.get("LINE")) == "29" and clean_text(row.get("LSTUDY")) == "4"

    students_map = load_csv_column(Config.IPEDS_EF2024A_PATH, "EFTOTLT", parse_int, row_filter=ef_total_filter)
    dorm_map = load_csv_column(Config.IPEDS_IC2024_PATH, "ROOMCAP", parse_int)
    accept_map = load_csv_column(Config.IPEDS_DRVADM2024_PATH, "ADM_RATE", parse_float)

    if hd_exists:
        with open(Config.IPEDS_HD2024_PATH, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                unitid = clean_text(row.get("UNITID"))
                if not unitid:
                    continue
                updates = {
                    "unitid": unitid,
                    "institution_id": unitid,
                    "ipeds_unit_ids": unitid,
                    "location_name": clean_text(row.get("INSTNM")),
                    "alias": clean_text(row.get("IALIAS")),
                    "street": clean_text(row.get("ADDR")),
                    "address": clean_text(row.get("ADDR")),
                    "city": clean_text(row.get("CITY")),
                    "state": clean_text(row.get("STABBR")),
                    "zip": clean_text(row.get("ZIP")),
                    "general_phone": clean_text(row.get("GENTELE")),
                    "telephone": clean_text(row.get("GENTELE")),
                    "ein": clean_text(row.get("EIN")),
                    "ope_id": clean_text(row.get("OPEID")),
                    "control": map_control(row.get("CONTROL")),
                    "institution_level": map_level(row.get("ICLEVEL")),
                    "highest_offering": map_highest_offering(row.get("HLOFFER")),
                    "ug_offering": map_offer(row.get("UGOFFER"), "undergraduate programs"),
                    "grad_offering": map_offer(row.get("GROFFER"), "graduate programs"),
                    "degree_granting_status": map_degree_grant(row.get("DEGGRANT")),
                    "locale": map_locale(row.get("LOCALE")),
                    "public_status": map_public_status(row.get("OPENPUBL")),
                    "post_secondary_status": map_post_secondary(row.get("POSTSEC")),
                    "fips_state_code": clean_text(row.get("FIPS")),
                    "fips_county_code": clean_text(row.get("COUNTYCD")),
                    "county": clean_text(row.get("COUNTYNM")),
                    "congressional_district": clean_text(row.get("CNGDSTCD")),
                    "website": clean_text(row.get("WEBADDR")),
                    "latitude": parse_float(row.get("LATITUDE")),
                    "longitude": parse_float(row.get("LONGITUD")),
                }
                if unitid in students_map:
                    updates["students_total"] = students_map[unitid]
                if unitid in dorm_map:
                    updates["dorm_capacity"] = dorm_map[unitid]
                if unitid in accept_map:
                    updates["acceptance_rate"] = accept_map[unitid]
                ensure_norm_fields(updates)
                match_kind, match_id = match_existing(updates)
                if match_kind == "existing":
                    merge_record(records[match_id], updates, overwrite=True)
                elif match_kind == "new":
                    merge_record(new_records[match_id], updates, overwrite=True)
                else:
                    new_records.append(dict(updates))
                    new_index = len(new_records) - 1
                    if updates.get("ope_id"):
                        new_by_ope.setdefault(clean_text(updates.get("ope_id")), new_index)
                    if updates.get("institution_id"):
                        new_by_ipeds.setdefault(clean_text(updates.get("institution_id")), new_index)
                    key = name_key(clean_text(updates.get("location_name")), clean_text(updates.get("state")), clean_text(updates.get("city")))
                    if key[0]:
                        new_by_name.setdefault(key, new_index)

    table_info = conn.execute("PRAGMA table_info(institutions)").fetchall()
    cols = [row[1] for row in table_info if row[1] != "id"]
    if records:
        update_rows = []
        for rec_id, rec in records.items():
            ensure_norm_fields(rec)
            update_rows.append([rec.get(c) for c in cols] + [rec_id])
        conn.executemany(
            f"UPDATE institutions SET {', '.join([f'{c}=?' for c in cols])} WHERE id=?",
            update_rows,
        )

    if new_records:
        for rec in new_records:
            ensure_norm_fields(rec)
        insert_rows = [[rec.get(c) for c in cols] for rec in new_records]
        placeholders = ",".join("?" for _ in cols)
        conn.executemany(
            f"INSERT INTO institutions({', '.join(cols)}) VALUES({placeholders})",
            insert_rows,
        )

    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_accredited_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (acc_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_ipeds_hd2024_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (hd_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_ipeds_ef2024a_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (ef_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_ipeds_ic2024_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (ic_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_ipeds_drvadm2024_mtime', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (drv_mtime,),
    )
    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('institutions_source', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (source_tag,),
    )
    chapter_info = conn.execute("PRAGMA table_info(chapters)").fetchall()
    if chapter_info:
        current_link_max = conn.execute("SELECT COALESCE(MAX(id), 0) FROM chapters").fetchone()[0]
        conn.execute(
            "INSERT INTO app_meta(key, value) VALUES('chapters_link_max_id', ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (str(int(current_link_max or 0)),),
        )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_name ON institutions(location_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_state ON institutions(state)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_students ON institutions(students_total)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_control ON institutions(control)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inst_unitid ON institutions(unitid)")

    _refresh_chapter_institution_links(conn)

    conn.commit()

def ensure_chapters_table(conn: sqlite3.Connection, bootstrap_related: bool = True) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )

    def table_columns(table_name: str) -> list[str]:
        return [row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]

    raw_cols = table_columns("chapters")
    if raw_cols and "chapter_uid" not in raw_cols and ("source_file" in raw_cols or "row_number" in raw_cols):
        conn.execute("ALTER TABLE chapters RENAME TO chapters_raw")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS chapters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chapter_uid TEXT UNIQUE,
            institution_id INTEGER,
            chapter_name TEXT,
            organization TEXT,
            school TEXT,
            city TEXT,
            state TEXT,
            instagram TEXT,
            chapter_id TEXT,
            status TEXT,
            founded_year INTEGER,
            notes TEXT,
            org_code TEXT,
            entity_type TEXT,
            scope TEXT,
            source_file TEXT,
            row_number TEXT
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_inst ON chapters(institution_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_org ON chapters(organization)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_uid ON chapters(chapter_uid)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_name ON chapters(chapter_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_school ON chapters(school)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chapter_state_city ON chapters(state, city)")

    raw_exists = bool(table_columns("chapters_raw"))
    if not raw_exists:
        return

    raw_count = conn.execute("SELECT COUNT(*) FROM chapters_raw").fetchone()[0]
    prev_count_row = conn.execute("SELECT value FROM app_meta WHERE key='chapters_raw_count'").fetchone()
    prev_count = int(prev_count_row[0]) if prev_count_row and str(prev_count_row[0]).isdigit() else -1
    existing_count = conn.execute("SELECT COUNT(*) FROM chapters").fetchone()[0]
    if existing_count > 0 and raw_count == prev_count:
        if bootstrap_related:
            _refresh_chapter_institution_links(conn)
            conn.commit()
        return

    if not bootstrap_related:
        return

    ensure_institutions_table(conn)
    inst_match_rows = conn.execute(
        "SELECT id, location_name, alias, parent_name, city, state FROM institutions"
    ).fetchall()
    match_index = _build_institution_match_index(inst_match_rows)

    conn.execute("DELETE FROM chapters")
    data_columns = [c for c in table_columns("chapters_raw") if c not in {"id"}]
    select_sql = "SELECT " + ", ".join([f'\"{c}\"' for c in data_columns]) + " FROM chapters_raw"
    chapter_rows = conn.execute(select_sql).fetchall()
    batch = []
    for r in chapter_rows:
        source_file = clean_text(r["source_file"]) if "source_file" in r.keys() else ""
        row_number = clean_text(r["row_number"]) if "row_number" in r.keys() else ""
        org_code, org_name, entity_type, scope = parse_meta_from_file(source_file)

        values: list[str] = []
        for c in data_columns:
            if c in {"source_file", "row_number"}:
                continue
            v = clean_text(r[c])
            if v and v not in {"[", "]"} and not v.lower().startswith("http"):
                values.append(v)

        if not values:
            continue

        chapter_name = detect_chapter(values)
        chapter_id = detect_chapter_id(values)
        school = detect_school(values)
        founded_year = detect_year(values)
        status = detect_status(values)
        notes = detect_notes(values)

        city = ""
        state = ""
        for v in values:
            if v in {chapter_name, school}:
                continue
            loc_city, loc_state = parse_location(v)
            city = city or loc_city
            state = state or loc_state

        if not any([chapter_name, school, city, state, status, founded_year]):
            continue

        chapter_uid = f"{source_file}::{row_number}" if source_file or row_number else chapter_id or chapter_name
        inst_id = _match_chapter_to_institution(school, state, city, match_index) if school else None

        batch.append(
            (
                chapter_uid,
                inst_id,
                chapter_name,
                org_name,
                school,
                city,
                state,
                "",
                chapter_id,
                status,
                int(founded_year) if str(founded_year).isdigit() else None,
                notes,
                org_code,
                entity_type,
                scope,
                source_file,
                row_number,
            )
        )

    if batch:
        conn.executemany(
            """
            INSERT INTO chapters
            (chapter_uid, institution_id, chapter_name, organization, school, city, state, instagram, chapter_id,
             status, founded_year, notes, org_code, entity_type, scope, source_file, row_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )

    conn.execute(
        "INSERT INTO app_meta(key, value) VALUES('chapters_raw_count', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (str(raw_count),),
    )
    _refresh_chapter_institution_links(conn)
    conn.commit()

def load_vendor_lookup(conn: sqlite3.Connection) -> tuple:
    rows = conn.execute(
        "SELECT vendor, organization, website, city, state, category, email, organization_norm, state_norm FROM vendors"
    ).fetchall()

    def org_keys(value: str) -> list[str]:
        raw = clean_text(value)
        if not raw:
            return []
        parts = [p.strip() for p in re.split(r",|/|&|\band\b", raw, flags=re.I) if clean_text(p)]
        keys = {norm_org(p) for p in parts if norm_org(p)}
        whole = norm_org(raw)
        if whole:
            keys.add(whole)
        lower_raw = raw.lower()
        for full_name, _kind in Config.ORG_MAP.values():
            if full_name.lower() in lower_raw:
                nk = norm_org(full_name)
                if nk:
                    keys.add(nk)
        return sorted(keys)

    exact_lookup = {}
    org_lookup = {}
    for r in rows:
        org_norms = org_keys(clean_text(r["organization"]))
        state_norm = clean_text(r["state_norm"])
        base = {
            "vendor": clean_text(r["vendor"]),
            "organization": clean_text(r["organization"]),
            "website": clean_text(r["website"]),
            "city": clean_text(r["city"]),
            "state": clean_text(r["state"]),
            "category": clean_text(r["category"]),
            "email": clean_text(r["email"]),
        }
        for org_norm in org_norms:
            if state_norm:
                rec_exact = dict(base)
                rec_exact["matchType"] = "Org + State"
                exact_lookup.setdefault((org_norm, state_norm), []).append(rec_exact)

            rec_org = dict(base)
            rec_org["matchType"] = "Org-wide"
            org_lookup.setdefault(org_norm, []).append(rec_org)

    return exact_lookup, org_lookup
